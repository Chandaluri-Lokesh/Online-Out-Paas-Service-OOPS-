from tkinter import *
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

class ApprovedRequestsPage:
    def __init__(self):
        self.approved_page = Tk()
        self.approved_page.geometry("600x400")
        self.approved_page.title("Approved Requests")

        self.approved_requests = self.load_approved_requests()

        self.current_request_index = 0

        self.request_label = Label(self.approved_page, text="", font=("Arial", 14))
        self.request_label.pack(pady=20)

        self.previous_button = Button(self.approved_page, text="Previous", font=("Arial", 12), width=10, command=self.show_previous_request)
        self.previous_button.pack(side=LEFT, padx=10)

        self.next_button = Button(self.approved_page, text="Next", font=("Arial", 12), width=10, command=self.show_next_request)
        self.next_button.pack(side=LEFT)

        self.delete_button = Button(self.approved_page, text="Delete", font=("Arial", 12), width=10, command=self.delete_request)
        self.delete_button.pack(side=LEFT, padx=10)

        self.update_request()

        self.approved_page.mainloop()

    def load_approved_requests(self):
        wb = load_workbook("user_details.xlsx")
        dayout_sheet = wb["DayOut"]
        long_leave_sheet = wb["LongLeave"]

        approved_requests = []

        # Load approved dayout requests
        for row in dayout_sheet.iter_rows(min_row=2, values_only=True):
            if row[5] == "Approved":
                approved_requests.append(f"Name: {row[0]}\nSubmission Date: {row[1]}\nOutTime: {row[2]}\nInTime: {row[3]}\nReason: {row[4]}\nStatus: {row[5]}\n")

        # Load approved long leave requests
        for row in long_leave_sheet.iter_rows(min_row=2, values_only=True):
            if row[7] == "Approved":
                approved_requests.append(f"Name: {row[0]}\nSubmission Date: {row[1]}\nFrom: {row[2]}\nTo: {row[3]}\nState: {row[4]}\nCity: {row[5]}\nReason: {row[6]}\nStatus: {row[7]}\n")

        return approved_requests

    def update_request(self):
        if self.current_request_index < len(self.approved_requests):
            request_text = self.approved_requests[self.current_request_index]
            self.request_label.config(text=request_text)
        else:
            self.request_label.config(text="No more approved requests")

    def show_previous_request(self):
        if self.current_request_index > 0:
            self.current_request_index -= 1
            self.update_request()

    def show_next_request(self):
        if self.current_request_index < len(self.approved_requests) - 1:
            self.current_request_index += 1
            self.update_request()

    def delete_request(self):
        if self.current_request_index < len(self.approved_requests):
            deleted_request = self.approved_requests.pop(self.current_request_index)
            self.update_request()

            # Save the updated approved requests back to the Excel file
            wb = load_workbook("user_details.xlsx")
            dayout_sheet = wb["dayout"]
            long_leave_sheet = wb["long_leave"]

            # Delete the request from the dayout sheet
            for row in dayout_sheet.iter_rows(min_row=2, values_only=True):
                if f"Name: {row[0]}\nSubmission Date: {row[1]}\nOutTime: {row[2]}\nInTime: {row[3]}\nReason: {row[4]}\nStatus: {row[5]}\n" == deleted_request:
                    dayout_sheet.delete_rows(row[0].row)
                    break

            # Delete the request from the long leave sheet
            for row in long_leave_sheet.iter_rows(min_row=2, values_only=True):
                if f"Name: {row[0]}\nSubmission Date: {row[1]}\nFrom: {row[2]}\nTo: {row[3]}\nState: {row[4]}\nCity: {row[5]}\nReason: {row[6]}\nStatus: {row[7]}\n" == deleted_request:
                    long_leave_sheet.delete_rows(row[0].row)
                    break

            wb.save("user_details.xlsx")

            messagebox.showinfo("Request Deleted", "The approved request has been deleted.")

approved_page = ApprovedRequestsPage()
