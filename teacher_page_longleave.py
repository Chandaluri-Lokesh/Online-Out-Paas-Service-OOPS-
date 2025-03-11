from tkinter import *
import tkinter as tk
from tkinter import messagebox
import subprocess
import openpyxl

class TeacherPage:
    def __init__(self):
        self.homepage = Tk()
        self.homepage.geometry("600x400")
        self.homepage.title("Teacher Page")

        self.workbook = openpyxl.load_workbook("user_details.xlsx")  # Replace with your actual Excel file path
        self.sheet = self.workbook["LongLeave"]

        self.requests = self.load_requests()

        self.current_request_index = 0

        self.status_var = StringVar()
        self.status_var.set("")

        self.request_label = Label(self.homepage, text="", font=("Arial", 14))
        self.request_label.pack(pady=20)

        self.status_label = Label(self.homepage, textvariable=self.status_var, font=("Arial", 14))
        self.status_label.pack()

        self.approve_button = Button(self.homepage, text="Approve", font=("Arial", 12), width=10, command=self.approve_request)
        self.approve_button.pack(pady=10)

        self.disapprove_button = Button(self.homepage, text="Decline", font=("Arial", 12), width=10, command=self.decline_request)
        self.disapprove_button.pack(pady=5)

        self.open_teacher_page_button = Button(self.homepage, text="Dayout Approval Page", font=("Arial", 12), width=15, command=self.open_teacher_page)
        self.open_teacher_page_button.place(x=10, y=375, anchor="sw")

        self.update_request()

        self.homepage.mainloop()

    def load_requests(self):
        requests = []
        for row in self.sheet.iter_rows(values_only=True):
            if len(row) >= 8:
                request = {
                    "Name": row[0],
                    "Submission Date": row[1],
                    "From": row[2],
                    "To": row[3],
                    "State": row[4],
                    "City": row[5],
                    "Reason": row[6],
                    "Status": row[7]
                }
                requests.append(request)
        return requests

    def update_request(self):
        while self.current_request_index < len(self.requests):
            request = self.requests[self.current_request_index]
            if request["Status"] == "":
                request_text = self.format_request_text(request)
                self.request_label.config(text=request_text)
                self.status_var.set("")
                return
            self.current_request_index += 1

        # No more requests or all requests have a status assigned
        self.request_label.config(text="No more requests")
        self.status_var.set("")

    def approve_request(self):
        if self.current_request_index < len(self.requests):
            request = self.requests[self.current_request_index]
            if request["Status"] == "":
                request["Status"] = "Approved"
                self.update_request()
                self.save_requests_to_excel()
                self.status_var.set("Request Approved")

    def decline_request(self):
        if self.current_request_index < len(self.requests):
            request = self.requests[self.current_request_index]
            if request["Status"] == "":
                request["Status"] = "Declined"
                self.update_request()
                self.save_requests_to_excel()
                self.status_var.set("Request Declined")

    def save_requests_to_excel(self):
        for row_index, request in enumerate(self.requests, start=2):
            self.sheet.cell(row=row_index, column=8).value = request["Status"]
        self.workbook.save("your_excel_file.xlsx")  # Replace with your actual Excel file path

    def format_request_text(self, request):
        return f"Name: {request['Name']}\nSubmission Date: {request['Submission Date']}\nFrom: {request['From']}\nTo: {request['To']}\nState: {request['State']}\nCity: {request['City']}\nReason: {request['Reason']}"

    def open_teacher_page(self):
        subprocess.Popen(["python", "teacher_page_dayout.py"])

TeacherPage()
