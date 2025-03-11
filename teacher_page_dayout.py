import tkinter as tk
from tkinter import Label, Button, StringVar
from openpyxl import load_workbook

class TeacherPage:
    def __init__(self):
        self.homepage = tk.Tk()
        self.homepage.geometry("600x400")
        self.homepage.title("Teacher Page")

        self.sheet_name = "DayOut"
        self.requests = self.load_requests()

        self.current_request_index = 0

        self.status_var = StringVar()
        self.status_var.set("")

        self.request_label = Label(self.homepage, text="", font=("Arial", 14))
        self.request_label.pack(pady=20)

        self.status_label = Label(self.homepage, textvariable=self.status_var, font=("Arial", 14))
        self.status_label.pack()

        self.approve_button = Button(self.homepage, text="Approve", font=("Arial", 12), width=10, command=self.approve_request)
        self.approve_button.pack(pady=10)

        self.disapprove_button = Button(self.homepage, text="Disapprove", font=("Arial", 12), width=10, command=self.disapprove_request)
        self.disapprove_button.pack(pady=5)

        self.open_teacher_page_button = Button(self.homepage, text="Long Leave Approval Page", font=("Arial", 12), width=15, command=self.open_teacher_page)
        self.open_teacher_page_button.place(x=10, y=375, anchor="sw")

        self.update_request()

        self.homepage.mainloop()

    def load_requests(self):
        workbook = load_workbook("user_details.xlsx")
        sheet = workbook["DayOut"]
        requests = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            request = {
                'row_index': row_index,
                'Name': row[0],
                'Submission Date': row[1],
                'OutTime': row[2],
                'InTime': row[3],
                'Reason': row[4],
                'Status': row[5]
            }
            requests.append(request)
        workbook.close()
        return requests

    def update_request(self):
        if self.current_request_index < len(self.requests):
            request = self.requests[self.current_request_index]
            self.request_label.config(text=f"InTime: {request['InTime']}\nOutTime: {request['OutTime']}\nReason: {request['Reason']}")
            self.status_var.set("")
        else:
            self.request_label.config(text="No more requests")
            self.status_var.set("")

    def update_excel(self):
        workbook = load_workbook("user_details.xlsx")
        sheet = workbook["DayOut"]
        for request in self.requests:
            row_index = request['row_index']
            status = request['Status']
            sheet.cell(row=row_index, column=6).value = status  # Update column index to 6 for "Status" column
        workbook.save("user_details.xlsx")
        workbook.close()
        
    def approve_request(self):
        if self.current_request_index < len(self.requests):
            self.requests[self.current_request_index]['Status'] = 'Approved'
            self.update_excel()
            self.status_var.set("Request Approved")
            self.current_request_index += 1
            self.update_request()

    def disapprove_request(self):
        if self.current_request_index < len(self.requests):
            self.requests[self.current_request_index]['Status'] = 'Disapproved'
            self.update_excel()
            self.status_var.set("Request Disapproved")
            self.current_request_index += 1
            self.update_request()

    def open_teacher_page(self):
        pass

TeacherPage()
